import io
import os
import zipfile
from dataclasses import dataclass, asdict
from typing import Any, Dict, List, Optional, Tuple

import cv2
import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
from PIL import Image, ImageSequence
from skimage.metrics import structural_similarity as ssim


TIMEOUT = 40
USER_AGENT = "photo-report-checker/1.0"


@dataclass
class MatchResult:
    row_number: int
    photo_id: str
    photo_url: str
    status: str
    confidence: float
    best_frame_index: int
    total_frames_checked: int
    best_method: str
    good_matches: int
    inliers: int
    inlier_ratio: float
    matched_area_ratio: float
    ssim_score: float
    hist_score: float
    raw_score: float
    error: str = ""


def pil_to_bgr(img: Image.Image) -> np.ndarray:
    rgb = img.convert("RGB")
    arr = np.array(rgb)
    return cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)


VIDEO_EXTENSIONS = {".mp4", ".mov", ".avi", ".webm", ".mkv"}


def load_creative_frames(path: str, gif_frame_step: int = 4, max_frames: int = 24) -> List[np.ndarray]:
    ext = os.path.splitext(path)[1].lower()

    if ext in VIDEO_EXTENSIONS:
        cap = cv2.VideoCapture(path)
        frames: List[np.ndarray] = []
        i = 0
        used = 0
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            if i % max(1, gif_frame_step) == 0:
                frames.append(frame)
                used += 1
                if used >= max_frames:
                    break
            i += 1
        cap.release()
        return frames or []

    img = Image.open(path)
    if ext != ".gif":
        return [pil_to_bgr(img)]

    frames = []
    used = 0
    for i, frame in enumerate(ImageSequence.Iterator(img)):
        if i % max(1, gif_frame_step) != 0:
            continue
        frames.append(pil_to_bgr(frame.copy()))
        used += 1
        if used >= max_frames:
            break

    return frames or [pil_to_bgr(img)]


def read_photo_links_from_excel(path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row = None
    photo_col = None
    headers_map: Dict[int, str] = {}

    for row in range(1, min(ws.max_row, 50) + 1):
        values = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            values.append("" if value is None else str(value).strip())

        lowered = [v.lower() for v in values]
        if "фотоотчет" in lowered or "фотоотчёт" in lowered:
            header_row = row
            for idx, name in enumerate(values, start=1):
                if name:
                    headers_map[idx] = name
                    if name.lower().strip() in ("фотоотчет", "фотоотчёт"):
                        photo_col = idx
            break

    if header_row is None or photo_col is None:
        raise ValueError('Не удалось найти заголовок "Фотоотчет" в файле.')

    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row, photo_col)
        photo_id = "" if cell.value is None else str(cell.value).strip()
        photo_url = cell.hyperlink.target if cell.hyperlink else ""

        if not photo_id and not photo_url:
            continue

        row_data = {
            "row_number": row,
            "photo_id": photo_id,
            "photo_url": photo_url,
        }

        for col_idx, header in headers_map.items():
            if col_idx == photo_col:
                continue
            row_data[header] = ws.cell(row, col_idx).value

        rows.append(row_data)

    return rows


def download_image(url: str, session: requests.Session) -> Image.Image:
    resp = session.get(url, timeout=TIMEOUT)
    resp.raise_for_status()
    return Image.open(io.BytesIO(resp.content)).convert("RGB")


def resize_keep_aspect(img: np.ndarray, max_side: int) -> np.ndarray:
    h, w = img.shape[:2]
    scale = min(max_side / max(h, w), 1.0)
    if scale == 1.0:
        return img
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    return cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_AREA)


def enhance_gray(img_bgr: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray = clahe.apply(gray)
    return gray


def compute_features(gray: np.ndarray, method: str):
    if method == "akaze":
        detector = cv2.AKAZE_create()
        norm = cv2.NORM_HAMMING
    elif method == "orb":
        detector = cv2.ORB_create(
            nfeatures=6000,
            scaleFactor=1.2,
            nlevels=8,
            edgeThreshold=15,
            patchSize=31,
            fastThreshold=5,
        )
        norm = cv2.NORM_HAMMING
    else:
        raise ValueError(f"Unknown method: {method}")

    kp, des = detector.detectAndCompute(gray, None)
    return kp, des, norm


def order_quad(pts: np.ndarray) -> np.ndarray:
    pts = np.asarray(pts, dtype=np.float32)
    s = pts.sum(axis=1)
    d = np.diff(pts, axis=1).reshape(-1)
    ordered = np.zeros((4, 2), dtype=np.float32)
    ordered[0] = pts[np.argmin(s)]
    ordered[2] = pts[np.argmax(s)]
    ordered[1] = pts[np.argmin(d)]
    ordered[3] = pts[np.argmax(d)]
    return ordered


def perspective_warp(img_bgr: np.ndarray, quad: np.ndarray, target_size: Tuple[int, int]) -> np.ndarray:
    quad = order_quad(quad)
    w, h = target_size
    dst = np.array([[0, 0], [w - 1, 0], [w - 1, h - 1], [0, h - 1]], dtype=np.float32)
    M = cv2.getPerspectiveTransform(quad, dst)
    return cv2.warpPerspective(img_bgr, M, (w, h))


def normalized_ssim(a_bgr: np.ndarray, b_bgr: np.ndarray) -> float:
    a = cv2.cvtColor(a_bgr, cv2.COLOR_BGR2GRAY)
    b = cv2.cvtColor(b_bgr, cv2.COLOR_BGR2GRAY)
    if a.shape != b.shape:
        b = cv2.resize(b, (a.shape[1], a.shape[0]), interpolation=cv2.INTER_AREA)
    score = ssim(a, b)
    if score != score:
        return 0.0
    return float(score)


def hist_similarity(a_bgr: np.ndarray, b_bgr: np.ndarray) -> float:
    a = cv2.cvtColor(a_bgr, cv2.COLOR_BGR2HSV)
    b = cv2.cvtColor(b_bgr, cv2.COLOR_BGR2HSV)

    if a.shape[:2] != b.shape[:2]:
        b = cv2.resize(b, (a.shape[1], a.shape[0]), interpolation=cv2.INTER_AREA)

    hist_a = cv2.calcHist([a], [0, 1], None, [32, 32], [0, 180, 0, 256])
    hist_b = cv2.calcHist([b], [0, 1], None, [32, 32], [0, 180, 0, 256])

    cv2.normalize(hist_a, hist_a)
    cv2.normalize(hist_b, hist_b)

    score = cv2.compareHist(hist_a, hist_b, cv2.HISTCMP_CORREL)
    score = max(-1.0, min(1.0, float(score)))
    return (score + 1.0) / 2.0


def draw_polygon(img_bgr: np.ndarray, quad: np.ndarray) -> np.ndarray:
    out = img_bgr.copy()
    pts = order_quad(quad).astype(np.int32).reshape((-1, 1, 2))
    cv2.polylines(out, [pts], True, (0, 255, 0), 4)
    return out


def try_match_one_method(creative_bgr: np.ndarray, photo_bgr: np.ndarray, method: str) -> Optional[Dict[str, Any]]:
    creative_bgr = resize_keep_aspect(creative_bgr, max_side=1000)
    photo_bgr = resize_keep_aspect(photo_bgr, max_side=1800)

    creative_gray = enhance_gray(creative_bgr)
    photo_gray = enhance_gray(photo_bgr)

    kp1, des1, norm = compute_features(creative_gray, method)
    kp2, des2, _ = compute_features(photo_gray, method)

    if des1 is None or des2 is None or len(kp1) < 8 or len(kp2) < 8:
        return None

    bf = cv2.BFMatcher(norm, crossCheck=False)
    raw_matches = bf.knnMatch(des1, des2, k=2)

    ratio = 0.78 if method == "akaze" else 0.75
    good = []
    for pair in raw_matches:
        if len(pair) != 2:
            continue
        m, n = pair
        if m.distance < ratio * n.distance:
            good.append(m)

    if len(good) < 6:
        return None

    src_pts = np.float32([kp1[m.queryIdx].pt for m in good]).reshape(-1, 1, 2)
    dst_pts = np.float32([kp2[m.trainIdx].pt for m in good]).reshape(-1, 1, 2)

    H, mask = cv2.findHomography(src_pts, dst_pts, cv2.RANSAC, 5.0)
    if H is None or mask is None:
        return None

    inliers = int(mask.ravel().sum())
    inlier_ratio = inliers / max(1, len(good))

    h1, w1 = creative_gray.shape[:2]
    corners = np.float32([[0, 0], [w1 - 1, 0], [w1 - 1, h1 - 1], [0, h1 - 1]]).reshape(-1, 1, 2)
    projected = cv2.perspectiveTransform(corners, H).reshape(-1, 2)

    photo_h, photo_w = photo_gray.shape[:2]
    poly_area = abs(cv2.contourArea(projected.astype(np.float32)))
    area_ratio = poly_area / float(max(1, photo_h * photo_w))
    if poly_area <= 0:
        return None

    try:
        warped = perspective_warp(photo_bgr, projected, (w1, h1))
    except Exception:
        return None

    ssim_score = normalized_ssim(creative_bgr, warped)
    hist_score = hist_similarity(creative_bgr, warped)

    raw_score = (
        0.22 * min(1.0, len(good) / 30.0)
        + 0.26 * min(1.0, inliers / 18.0)
        + 0.16 * min(1.0, inlier_ratio / 0.45)
        + 0.12 * min(1.0, area_ratio / 0.02)
        + 0.14 * ssim_score
        + 0.10 * hist_score
    )

    return {
        "method": method,
        "good_matches": len(good),
        "inliers": inliers,
        "inlier_ratio": float(inlier_ratio),
        "matched_area_ratio": float(area_ratio),
        "ssim_score": float(ssim_score),
        "hist_score": float(hist_score),
        "raw_score": float(raw_score),
        "projected": projected,
        "warped": warped,
        "photo_resized": photo_bgr,
        "creative_resized": creative_bgr,
    }


def classify_candidate(c: Dict[str, Any]) -> Tuple[str, float]:
    good = c["good_matches"]
    inliers = c["inliers"]
    inlier_ratio = c["inlier_ratio"]
    area_ratio = c["matched_area_ratio"]
    ssim_score = c["ssim_score"]
    hist_score = c["hist_score"]
    raw_score = c["raw_score"]

    yes_rule_1 = (
        inliers >= 6 and inlier_ratio >= 0.13 and area_ratio >= 0.001 and (ssim_score >= 0.20 or hist_score >= 0.52)
    )
    yes_rule_2 = good >= 10 and inliers >= 8 and area_ratio >= 0.0015
    yes_rule_3 = raw_score >= 0.40 and inliers >= 6

    review_rule = (inliers >= 4 and area_ratio >= 0.0008) or raw_score >= 0.28

    confidence = min(0.99, max(0.0, 0.25 + raw_score))

    if yes_rule_1 or yes_rule_2 or yes_rule_3:
        return "YES", confidence
    if review_rule:
        return "REVIEW", min(confidence, 0.84)
    return "NO", min(confidence, 0.45)


def score_match(creative_bgr: np.ndarray, photo_bgr: np.ndarray) -> Dict[str, Any]:
    candidates = []
    for method in ("akaze", "orb"):
        cand = try_match_one_method(creative_bgr, photo_bgr, method)
        if cand is not None:
            status, confidence = classify_candidate(cand)
            cand["status"] = status
            cand["confidence"] = confidence
            candidates.append(cand)

    if not candidates:
        return {
            "status": "NO",
            "confidence": 0.05,
            "best_method": "",
            "good_matches": 0,
            "inliers": 0,
            "inlier_ratio": 0.0,
            "matched_area_ratio": 0.0,
            "ssim_score": 0.0,
            "hist_score": 0.0,
            "raw_score": 0.0,
            "projected": None,
            "warped": None,
            "photo_resized": None,
            "creative_resized": None,
        }

    rank = {"YES": 3, "REVIEW": 2, "NO": 1}
    candidates.sort(
        key=lambda c: (rank[c["status"]], c["confidence"], c["raw_score"], c["ssim_score"], c["inliers"]),
        reverse=True,
    )
    best = candidates[0]
    best["best_method"] = best["method"]
    return best


def best_match_across_frames(creative_frames: List[np.ndarray], photo_bgr: np.ndarray) -> Dict[str, Any]:
    best = None
    rank = {"YES": 3, "REVIEW": 2, "NO": 1}

    for idx, frame in enumerate(creative_frames):
        cur = score_match(frame, photo_bgr)
        cur["frame_index"] = idx

        if best is None:
            best = cur
            continue

        cur_key = (rank[cur["status"]], cur["confidence"], cur["raw_score"], cur["ssim_score"], cur["inliers"])
        best_key = (rank[best["status"]], best["confidence"], best["raw_score"], best["ssim_score"], best["inliers"])
        if cur_key > best_key:
            best = cur

    return best


def save_debug(debug_dir: str, row_number: int, photo_id: str, best: Dict[str, Any]) -> None:
    os.makedirs(debug_dir, exist_ok=True)
    safe_id = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(photo_id or row_number))
    prefix = f"row_{row_number}_{safe_id}"

    if best.get("projected") is not None and best.get("photo_resized") is not None:
        outlined = draw_polygon(best["photo_resized"], best["projected"])
        cv2.imwrite(os.path.join(debug_dir, f"{prefix}_outlined.jpg"), outlined)

    if best.get("warped") is not None:
        cv2.imwrite(os.path.join(debug_dir, f"{prefix}_warped.jpg"), best["warped"])

    if best.get("creative_resized") is not None:
        cv2.imwrite(os.path.join(debug_dir, f"{prefix}_creative.jpg"), best["creative_resized"])


def zip_folder(folder_path: str, zip_path: str) -> None:
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(folder_path):
            for file in files:
                full = os.path.join(root, file)
                arc = os.path.relpath(full, folder_path)
                zf.write(full, arc)


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT})
    return s


def save_results(rows: List[Dict[str, Any]], path: str) -> None:
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="results")
        if "status" in df.columns:
            review_df = df[df["status"].isin(["REVIEW", "NO"])].copy()
            review_df.to_excel(writer, index=False, sheet_name="needs_review")


def run_check(creative_path: str, report_path: str, out_dir: str) -> Tuple[str, Optional[str], str]:
    """Sync entry point — run in a thread executor. Returns (result_xlsx, debug_zip_or_None, summary)."""
    creative_frames = load_creative_frames(creative_path)
    report_rows = read_photo_links_from_excel(report_path)
    total = len(report_rows)
    total_frames = len(creative_frames)

    session = make_session()
    out_rows: List[Dict[str, Any]] = []
    debug_dir = os.path.join(out_dir, "debug")

    for row in report_rows:
        base: Dict[str, Any] = {
            "row_number": row["row_number"],
            "photo_id": row["photo_id"],
            "photo_url": row["photo_url"],
        }
        for k, v in row.items():
            if k not in base:
                base[k] = v

        try:
            if not row["photo_url"]:
                res = MatchResult(
                    row_number=row["row_number"],
                    photo_id=row["photo_id"],
                    photo_url=row["photo_url"],
                    status="REVIEW",
                    confidence=0.0,
                    best_frame_index=-1,
                    total_frames_checked=total_frames,
                    best_method="",
                    good_matches=0,
                    inliers=0,
                    inlier_ratio=0.0,
                    matched_area_ratio=0.0,
                    ssim_score=0.0,
                    hist_score=0.0,
                    raw_score=0.0,
                    error="Нет ссылки на фото",
                )
            else:
                photo_img = download_image(row["photo_url"], session)
                photo_bgr = pil_to_bgr(photo_img)
                best = best_match_across_frames(creative_frames, photo_bgr)

                if best["status"] in ("REVIEW", "NO"):
                    save_debug(debug_dir, row["row_number"], row["photo_id"], best)

                res = MatchResult(
                    row_number=row["row_number"],
                    photo_id=row["photo_id"],
                    photo_url=row["photo_url"],
                    status=best["status"],
                    confidence=round(best["confidence"], 4),
                    best_frame_index=int(best["frame_index"]),
                    total_frames_checked=total_frames,
                    best_method=str(best.get("best_method", best.get("method", ""))),
                    good_matches=int(best["good_matches"]),
                    inliers=int(best["inliers"]),
                    inlier_ratio=round(best["inlier_ratio"], 4),
                    matched_area_ratio=round(best["matched_area_ratio"], 6),
                    ssim_score=round(best["ssim_score"], 4),
                    hist_score=round(best["hist_score"], 4),
                    raw_score=round(best["raw_score"], 4),
                )
        except Exception as e:
            res = MatchResult(
                row_number=row["row_number"],
                photo_id=row["photo_id"],
                photo_url=row["photo_url"],
                status="REVIEW",
                confidence=0.0,
                best_frame_index=-1,
                total_frames_checked=total_frames,
                best_method="",
                good_matches=0,
                inliers=0,
                inlier_ratio=0.0,
                matched_area_ratio=0.0,
                ssim_score=0.0,
                hist_score=0.0,
                raw_score=0.0,
                error=f"{type(e).__name__}: {e}",
            )

        merged = dict(base)
        merged.update(asdict(res))
        out_rows.append(merged)

    result_path = os.path.join(out_dir, "result.xlsx")
    save_results(out_rows, result_path)

    debug_zip: Optional[str] = None
    if os.path.exists(debug_dir):
        debug_zip = os.path.join(out_dir, "debug.zip")
        zip_folder(debug_dir, debug_zip)

    df = pd.DataFrame(out_rows)
    yes_count = int((df["status"] == "YES").sum()) if "status" in df.columns else 0
    review_count = int((df["status"] == "REVIEW").sum()) if "status" in df.columns else 0
    no_count = int((df["status"] == "NO").sum()) if "status" in df.columns else 0

    summary = (
        f"Проверка завершена.\n\n"
        f"Всего строк: {total}\n"
        f"YES (совпадает): {yes_count}\n"
        f"REVIEW (нужна ручная проверка): {review_count}\n"
        f"NO (не совпадает): {no_count}"
    )

    return result_path, debug_zip, summary


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}


def run_check_from_zip(creative_path: str, zip_path: str, out_dir: str) -> Tuple[str, Optional[str], str]:
    """Like run_check but reads photos from a ZIP archive instead of downloading from URLs."""
    creative_frames = load_creative_frames(creative_path)
    total_frames = len(creative_frames)

    extract_dir = os.path.join(out_dir, "extracted")
    os.makedirs(extract_dir, exist_ok=True)

    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(extract_dir)

    photo_paths: List[Tuple[int, str]] = []
    for root, _, files in os.walk(extract_dir):
        for fname in sorted(files):
            if os.path.splitext(fname)[1].lower() in IMAGE_EXTENSIONS:
                photo_paths.append((len(photo_paths) + 1, os.path.join(root, fname)))

    total = len(photo_paths)
    out_rows: List[Dict[str, Any]] = []
    debug_dir = os.path.join(out_dir, "debug")

    for row_number, photo_path in photo_paths:
        filename = os.path.relpath(photo_path, extract_dir)
        try:
            photo_img = Image.open(photo_path).convert("RGB")
            photo_bgr = pil_to_bgr(photo_img)
            best = best_match_across_frames(creative_frames, photo_bgr)

            if best["status"] in ("REVIEW", "NO"):
                save_debug(debug_dir, row_number, filename, best)

            res = MatchResult(
                row_number=row_number,
                photo_id=filename,
                photo_url="",
                status=best["status"],
                confidence=round(best["confidence"], 4),
                best_frame_index=int(best["frame_index"]),
                total_frames_checked=total_frames,
                best_method=str(best.get("best_method", best.get("method", ""))),
                good_matches=int(best["good_matches"]),
                inliers=int(best["inliers"]),
                inlier_ratio=round(best["inlier_ratio"], 4),
                matched_area_ratio=round(best["matched_area_ratio"], 6),
                ssim_score=round(best["ssim_score"], 4),
                hist_score=round(best["hist_score"], 4),
                raw_score=round(best["raw_score"], 4),
            )
        except Exception as e:
            res = MatchResult(
                row_number=row_number,
                photo_id=filename,
                photo_url="",
                status="REVIEW",
                confidence=0.0,
                best_frame_index=-1,
                total_frames_checked=total_frames,
                best_method="",
                good_matches=0,
                inliers=0,
                inlier_ratio=0.0,
                matched_area_ratio=0.0,
                ssim_score=0.0,
                hist_score=0.0,
                raw_score=0.0,
                error=f"{type(e).__name__}: {e}",
            )

        out_rows.append(asdict(res))

    result_path = os.path.join(out_dir, "result.xlsx")
    save_results(out_rows, result_path)

    debug_zip: Optional[str] = None
    if os.path.exists(debug_dir):
        debug_zip = os.path.join(out_dir, "debug.zip")
        zip_folder(debug_dir, debug_zip)

    df = pd.DataFrame(out_rows)
    yes_count = int((df["status"] == "YES").sum()) if "status" in df.columns else 0
    review_count = int((df["status"] == "REVIEW").sum()) if "status" in df.columns else 0
    no_count = int((df["status"] == "NO").sum()) if "status" in df.columns else 0

    summary = (
        f"Проверка завершена.\n\n"
        f"Всего фото: {total}\n"
        f"YES (совпадает): {yes_count}\n"
        f"REVIEW (нужна ручная проверка): {review_count}\n"
        f"NO (не совпадает): {no_count}"
    )

    return result_path, debug_zip, summary
